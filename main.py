import string
import random
import openpyxl
import os
from openpyxl import load_workbook

string.ascii_letters = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz1234567890!?@$#"§%&/()={[]}<>^'
cwd = os.getcwd()

def main(length, application):
    password = generatePw(length)
    print(password)
    choose = input("You want to keep this password (y/n): ")
    if choose == "y" or choose == "Y":
        setXlsxData(application, password)
    else:
        main(length, password)

def generatePw(length):
    password = ""
    for x in range(int(length)):
        password += random.choice(string.ascii_letters)
    return password

def getXlsxData():
    path = cwd + "\\password_list.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    maxRow = sheet_obj.max_row
    return maxRow

def setXlsxData(application, password):
    wb = load_workbook("password_list.xlsx")
    sheet = wb.active
    nextRow = int(getXlsxData())+1
    columns = sheet["A" + str(nextRow)].value = application
    columns = sheet["B" + str(nextRow)].value = password
    wb.save("password_list.xlsx")



application = input("Application: ")
length = input("Length: ")
main(length, application)